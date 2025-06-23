import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils import timezone

def export_products_csv(products):
    """Экспорт товаров в CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="products_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Название', 'Описание', 'Цена', 'Количество'])
    
    for product in products:
        writer.writerow([
            product.id,
            product.name,
            product.description,
            product.price,
            product.stock_quantity
        ])
    
    return response

def export_products_excel(products):
    """Экспорт товаров в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    
    # Заголовки
    headers = ['ID', 'Название', 'Описание', 'Цена', 'Количество']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
    
    # Данные
    for row_num, product in enumerate(products, 2):
        ws[f'A{row_num}'] = product.id
        ws[f'B{row_num}'] = product.name
        ws[f'C{row_num}'] = product.description
        ws[f'D{row_num}'] = float(product.price)
        ws[f'E{row_num}'] = product.stock_quantity
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="products_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

def get_low_stock_products(threshold=10):
    """Получить товары с низким запасом"""
    from .models import Product
    return Product.objects.filter(stock_quantity__lte=threshold)

def get_dashboard_stats():
    """Получить расширенную статистику для главной страницы"""
    from .models import Product, Order, Customer, InventoryLevel
    from django.db.models import Avg, Sum, Count
    
    stats = {}
    
    # Базовая статистика
    stats['total_products'] = Product.objects.count()
    stats['total_customers'] = Customer.objects.count()
    stats['active_orders'] = Order.objects.filter(status='active').count()
    stats['total_orders'] = Order.objects.count()
    
    # Финансовая статистика
    stats['total_sales'] = Order.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    stats['average_order_value'] = Order.objects.aggregate(avg=Avg('total_amount'))['avg'] or 0
    
    # Статистика по товарам
    stats['average_product_price'] = Product.objects.aggregate(avg=Avg('price'))['avg'] or 0
    stats['total_stock_value'] = sum(p.price * p.stock_quantity for p in Product.objects.all())
    
    # Предупреждения
    stats['low_stock_products'] = get_low_stock_products()
    stats['low_stock_count'] = stats['low_stock_products'].count()
    
    # Статистика заказов по статусам
    stats['orders_by_status'] = Order.objects.values('status').annotate(count=Count('id'))
    
    return stats
